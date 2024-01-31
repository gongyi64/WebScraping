# pySimpleGUI Version---Emily 自動ログインツール　202401
#Emilyへ自動ログインして、各自のメニューにアクセスする。案件制作を自動で。Ver.作成中。



import PySimpleGUI as sg
import re
import pandas as pd
import openpyxl

manNos = ('827861 相庭直史','406239 白川公一','380672 三角和浩','378035 岩田貴夫','805519 仲本祥子','880079 砂川航輝','806185 辻ひかる','880334   山城実咲','880518 古波蔵晃久','410993 黒岩英次','710463 山城徳松')

#manNos = ('827861 相庭直史　ap827861','406239 白川公一　ap406239','380672 三角和浩　ap380672','378035 岩田貴夫　ap378035','805519 仲本祥子   ap805519','880079 砂川航輝　ap880079','806185 辻ひかる　ap806185','880334   山城実咲 ap880334','880518 古波蔵晃久 ap880518','410993 黒岩英次   ap410993','710463 山城徳松   ap710463')

#manNos = ('827861','406239','380672','378035','805519','880079','806185','880334','880518','410993','710463')

sg.theme('Python')

layout =[[sg.Text('[NT_Emily_自動操作ソフト]',font = ('Noto Serif CJK JP',14))],

         [sg.Text('[Emilyに別ブラウザでログインしたいときに.誰でログインしますか？] ',font = ('meiryo',10))],

         [sg.Listbox(manNos,size =(25,len(manNos)),key='-MN-')],

         [sg.Text('Text', key = '-text1-')],

         [sg.Button('実行', button_color=('red','#808080'),key = '-SUBMIT-')]]

window = sg.Window('Emily_APP',layout,size = (500,350))



while True:
    event,values = window.read()

    if event == sg.WIN_CLOSED:
        break

    elif event == '-SUBMIT-':
        window['-text1-'].update(values['-MN-'][0])
        input_eplyNo = values['-MN-'][0]


window.close()



# input_eplyNo = values['-MN-'][0]

print(input_eplyNo)#所得したのは、リスト型

eplyNo = re.findall(r'\d+', input_eplyNo)#名前を除去して社員番号のみにして、ログインに使用する。
eplyName = re.sub(r"[0-9]+", "", input_eplyNo)#社員番号削除して氏名のみに。
# eplyNo = eplyNo[:7]

print(eplyNo[0])#リストの要素を文字列として取得。この場合は、要素１つなので[0]
print(eplyName)

# login_name = str(eplyName[0])
#
# print(login_name)
#eplypwd = re.findall(r'^[a-zA-Z0-9]{7}$, input_eplyNo)#名前とマンナンバーを除去してPWDのみにして、ログインに使用する。7桁の英数字想定。ｓ+マンナンバーなど。

# selenium 4

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
import os
import time
import datetime
import sys
from selenium.webdriver.common.action_chains import ActionChains

import calendar
from selenium.webdriver.support.ui import Select


from selenium.webdriver.common.by import By

from selenium.webdriver.common.keys import Keys

time.sleep(3)
#from webdriver_manager.chrome import ChromeDriverManager

#driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

#最新のドライバーだとエラーが出るのでその対応策。https://qiita.com/hs2023/questions/ffab105c5692692624ab

options = webdriver.ChromeOptions()


print("========== Emily　ログイン中========== ")
sg.popup_ok(f'Emilyへ{eplyName}でログインします',title = 'LOGIN')


# service = Service(driver_path)
# service = Service(executable_path=driver_path)# 2) executable_pathを指定
# driver = webdriver.Chrome(service=service)# 3) serviceを渡す

# 起動時にオプションをつける。（ポート指定により、起動済みのブラウザのドライバーを取得）
# driver_path = "C:\\Users\\406239\\AppData\\Local\\Programs\\Python\\Python39\\chromedriver_binary\\chromedriver.exe"

driver_path = "C:\\Users\\406239\\PycharmProjects\\pythonProject1\\chromedriver_binary\\chromedriver.exe"

#2023_11_09 chromdriver　118→119　更新

driver = webdriver.Chrome(service=ChromeService(driver_path))
# options.add_experimental_option("detach",True)#ドライバプロセスの終了後も開いたままにする。使い勝手よくなさそうなのでやめ。

options.add_experimental_option("debuggerAddress", "127.0.0.1:9333")
# driver = webdriver.Chrome(executable_path=driver_path, options=options)

# ページのタイトルを表示する

driver.get("https://test9.emily.nhk-tech.co.jp/GRANDIT/CM_AC_03_S01.aspx")

driver.maximize_window()

time.sleep(2)

driver.implicitly_wait(2)

#=====================================================================pwdを外ファイルからゲットルーチンここから。20230118実装

#Emily_Pass.xlsxがパスワード保管ファイル

file_name = 'c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/★勤務確認などのダウンロードデータ★/Emily_Files/Emily_Pass.xlsx'


df = pd.read_excel(file_name)

print(eplyNo[0])

eplyNo = int(eplyNo[0])#eplyNoは、リストなので、0番目を抽出し、intに変更。strだとエラー。

print (df[df['社員番号'] == eplyNo])#ログインする人の番号が含まれるdfを抽出


df_login = df[df['社員番号'] == eplyNo]


pwd = df_login.iat[0,2]#そのパスワードのみを抽出

print(pwd)


#=====================================================================ここからEmilyログイン



form = driver.find_element(By.XPATH,'//*[@id="LoginAccountText"]')

form.send_keys(eplyNo)#最初に取得したマンナンバーを入力

form = driver.find_element(By.XPATH,'//*[@id="PassText"]')

# form.send_keys('09-Hdo9QDw ')
form.send_keys(pwd)


time.sleep(2)

form = driver.find_element(By.XPATH,'//*[@id="LoginButton"]/span')

form.click()


time.sleep(3)




#ログインまでは、出来ているこの後は、まだ。メニュー画面にログインまで。#c_12

handle_array = driver.window_handles

print("handle_arrayの表示配列最初と次")#windowshandleは2つ
print(handle_array[0])
print(handle_array[1])
# print(handle_array[2])


driver.switch_to.window(handle_array[1])

print(driver.current_url)

time.sleep(2)


# 要素を特定する


#elem = driver.find_elements(By.XPATH,'//*[@id="__pageIframe_01703644264780"]')

# elem = driver.find_elements(By.CSS_SELECTOR,'#__pageIframe_01703644347699')
#iFrameをDevelopersで検索してそのiFrameのXPATHをさがす


#
driver.switch_to.frame(0)#iFrameの最初に切り替え。１つしかないが、classが毎回変わるので、1番目（０）のiFrameに切り替えるということにした。


dropdown = driver.find_element(By.CSS_SELECTOR,'#MenuDropDownList')
print('釦を取得できたらあり')
print(dropdown)

select = Select(dropdown)

select.select_by_index(1)





time.sleep(3)

#M技管理メニュー操作

driver.find_element(By.XPATH, '/html/body').send_keys(Keys.ENTER)  # エンターを押して、次メニューに更新。

time.sleep(3)

driver.find_element(By.XPATH, '//*[@id="c_11"]').click()  # 個別案件/要員をクリック
time.sleep(2)

driver.find_element(By.XPATH,'//*[@id="Form1"]/table/tbody/tr/td/table[2]/tbody/tr/td[2]/div[3]/table/tbody/tr/td[1]/div/a[2]').click()  # 182_個別案件入力をクリック
# driver.find_element(By.XPATH,'//*[@id="Form1"]/table/tbody/tr/td/table[2]/tbody/tr/td[2]/div[3]/table/tbody/tr/td[1]/div/a[2]').click()
time.sleep(2)


while True:

    sg.theme('SystemDefault')

    layout = [[sg.Text('案件作成年月を入力',text_color='#FF0000',font =( 'meiryo,6')),sg.InputText(size = (10,2),key= '-YM-')],
          # [sg.Text('誰の案件？',text_color='#FF0000',font =( 'meiryo,8')),sg.InputText(size = (10,2),key= '-NM-')],
              [sg.Listbox(manNos,size =(25,len(manNos)),key='-NM-')],
              [sg.Text('Text', key = '-text1-')],
              [sg.Button('入力', button_color=('red', '#808080'), key='-SUBMIT-'),
               sg.Text('入力ボタンを押した後,Windowを閉じてください。\nここから手動操作したいとき（新規案件作成以外）は、\n何も入力せずにそのままWindowを閉じてください。', font=('Noto Serif CJK JP', 10))]]

    window = sg.Window('案件自動作成ツール', layout, size=(500, 500))

    while True:
        event,values = window.read()

        if event == sg.WIN_CLOSED:
            break

        elif event == '-SUBMIT-':
            ym = values['-YM-']
            window['-text1-'].update(values['-NM-'][0])
            input_eplyNo = values['-NM-'][0]
            eplyName = re.sub(r"[0-9]+", "", input_eplyNo)#社員番号と名前から社員番号削除してフルネームのみに。
            window.close()

    window.close()

    driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)#エンターを押して、次メニューに更新。

    driver.implicitly_wait(3)





    handle_array = driver.window_handles
    #
    # print("handle_arrayの表示配列最初と次")  # windowshandleは2つ
    # print(handle_array[0])
    # print(handle_array[1])
    # # print(handle_array[2])
    #
    driver.switch_to.window(handle_array[1])


#os.kill(driver.service.process.pid,signal.SIGTERM)#ブラウザが閉じるのを止める。開きっぱなしにする。
    time.sleep(2)

    driver.switch_to.frame(1)#iFrameの最初に切り替え。２つあるが、2番目（1）のiFrameに切り替える。２人目の場合ここのフレームが見つからない。要修正。20240131

    form = driver.find_element(By.XPATH,'//*[@id="ProgramBusinessCodeText"]')

    form.send_keys('2006101343')#番組業務番号の入力

    driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)

    time.sleep(3)
    dropdown = driver.find_element(By.XPATH, '//*[@id="MediaBusinessClassCodeList"]')  # 業務実施区分　沖縄　リスト92番目

    select = Select(dropdown)

    select.select_by_index(92)#沖縄事業所をドロップダウンで選択→変更してよいかと言われるのでアラート解除を下に追加。

    handle_array = driver.window_handles

# print("別ページに切り替えた後のhandle_arrayの表示配列最初と次")#windowshandleは2つ結局かわらす。
# print(handle_array[0])
# print(handle_array[1])
    driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新

    driver.switch_to.window(handle_array[1])

    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()

    driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新

    driver.switch_to.frame(1)#iFrameの最初に切り替え。２つあるが、2番目（1）のiFrameに切り替える。

    monthday = str(calendar.monthrange(2024,6)[1])

    ym = str(ym)



    # print(ym)

    monthday = calendar.monthrange(int((ym[:4])),int((ym[4:])))[1]

    print(monthday)



    ymd_s = ym[:4]+'/'+ym[4:]+'/'+'01'#月の最初の日付　0000/00/00の書式で

    ymd_l = ym[:4]+'/'+ym[4:]+'/'+str(monthday)#月の最終日の日付け　　0000/00/00の書式で

    print(ymd_s)

    print(ymd_l)

    form = driver.find_element(By.XPATH,'//*[@id="BaseDateText"]')

    form.clear()

    form.send_keys(ymd_s)


    driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)

    form = driver.find_element(By.XPATH,'//*[@id="ProgramDeptText"]')

    form.clear()

    time.sleep(2)

    driver.find_element(By.XPATH,'//*[@id="ProgramDeptText"]').send_keys('557030')#実施担当部門沖縄　557030の入力。初期値は発注の福岡553010になっているため
    driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)

    time.sleep(2)


    # form = driver.find_element(By.XPATH,'//*[@id="ProgramBusinessCodeText"]')
    #
    # form.send_keys('2006101343')#番組業務番号の入力
    #
    # driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)


    Project1_No = 'M3P0000095-0H'#要員費

    Project2_No = 'M3P0000095-0I'#諸経費

    form = driver.find_element(By.XPATH,'//*[@id="ProjCodeText1"]')

    # print(form)

    form.send_keys(Project1_No)#Project1 要員費番号を入力

    driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)

    time.sleep(2)

    form = driver.find_element(By.XPATH,'//*[@id="ProjCodeText2"]')

    # print(form)

    form.send_keys(Project2_No)#物品費番号を入力


    driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)#エンターを押して、次メニューに更新。
    #
    time.sleep(2)

    # os.kill(driver.service.process.pid,signal.SIGTERM)#ブラウザが閉じるのを止める。開きっぱなしにする。


    #driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)#エンターを押して、次メニューに更新。
    driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新。



    driver.implicitly_wait(2)

    #dropdown3 = driver.find_element(By.CSS_SELECTOR, '#AcceptFormDtl1List')  # 受注形態詳細１　沖縄事業所入力　リスト22番目
    dropdown3 = driver.find_element(By.XPATH,'//*[@id="AcceptFormDtl1List"]')
    select = Select(dropdown3)

    select.select_by_index(22)#沖縄事業所をドロップダウンで選択→プロジェクト入力で自動入力されるのでそのまま

    driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新。

    #dropdown4 = driver.find_element(By.CSS_SELECTOR, '#AcceptFormDtl2List')  # N/A入力
    dropdown4 = driver.find_element(By.XPATH,'//*[@id="AcceptFormDtl2List"]')
    select = Select(dropdown4)

    select.select_by_index(1)

    driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新。

    dropdown = driver.find_element(By.XPATH, '//*[@id="StationInoutTypeList"]')  # 福岡　NHK局内選択

    select = Select(dropdown)

    select.select_by_index(13)







    driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)#エンターを押して

    time.sleep(1)

    dropdown = driver.find_element(By.XPATH, '//*[@id="ResourceList"]')  # リソース未定なし

    select = Select(dropdown)

    select.select_by_index(1)


    driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新。

    # os.kill(driver.service.process.pid,signal.SIGTERM)#ブラウザが閉じるのを止める。開きっぱなしにする。


    driver.find_element(By.XPATH,'//*[@id="TabItem2"]/span').click()
    #driver.find_element(By.CSS_SELECTOR,'#TabItem2 > span').click()

    time.sleep(2)

    handle_array = driver.window_handles

    print("別ページに切り替えた後のhandle_arrayの表示配列最初と次")#windowshandleは2つ結局かわらす。
    print(handle_array[0])
    print(handle_array[1])


    driver.switch_to.window(handle_array[1])

    driver.switch_to.frame(1)#iFrameの最初に切り替え。２つあるが、2番目（1）のiFrameに切り替える。

    driver.find_element(By.XPATH,'//*[@id="SubtitleText"]').send_keys(ym[:4]+'年'+ym[4:]+'月 '+eplyName)#副題


    driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)#エンターを押して、次メニューに更新。

    time.sleep(1)

    dropdown = driver.find_element(By.XPATH,'//*[@id="InpOpeDtlList"]')#作業詳細

    select = Select(dropdown)

    select.select_by_index(28)

    time.sleep(2)


    form = driver.find_element(By.XPATH,'//*[@id="InpSttDateText"]')#開始日時　日

    form.send_keys(ymd_s)

    #driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新。

    driver.find_element(By.XPATH,'//*[@id="InpSttTimeText"]').send_keys('00:00')#開始日時　時


    #driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新。



    form = driver.find_element(By.XPATH,'//*[@id="InpEndDateText"]')#終了日時　日

    form.send_keys(ymd_l)

    #driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新。


    driver.find_element(By.XPATH,'//*[@id="InpEndTimeText"]').send_keys('00:00')#終了日時　

    #driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新。

    # alert = driver.switch_to.alert
    # print(alert.text)
    # alert.accept()

    driver.find_element(By.XPATH,'//*[@id="InpSttDateText"]').send_keys(ymd_s)#開始日時

    time.sleep(1)
    driver.find_element(By.XPATH,'//*[@id="InpOpeDtlCntText1"]').clear()#担当　担当者数　をクリア

    driver.find_element(By.XPATH,'//*[@id="InpOpeDtlCntText1"]').send_keys(1)#担当　担当者数　１を入力

    driver.find_element(By.XPATH,'/html/body').click()#エンターを押して、次メニューに更新。


    time.sleep(1)

    #driver.find_element(By.XPATH,'/html/body').send_keys(Keys.ENTER)#エンターを押して、次メニューに更新。

    # driver.find_element(By.XPATH,'//*[@id="RegistButton"]/span').click()#登録ボタン

    driver.find_element(By.CSS_SELECTOR,'#RegistButton > span').click()


    #os.kill(driver.service.process.pid,signal.SIGTERM)#ブラウザが閉じるのを止める。開きっぱなしにする。

    driver.find_element(By.XPATH,'//*[@id="UpdateButton"]/span').click()#更新ボタン


    handle_array = driver.window_handles

    # print("別ページに切り替えた後のhandle_arrayの表示配列最初と次")#windowshandleは2つ結局かわらす。
    # print(handle_array[0])
    # print(handle_array[1])

    driver.switch_to.window(handle_array[1])

    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()



    driver.switch_to.frame(1)#iFrameの最初に切り替え。２つあるが、2番目（1）のiFrameに切り替える。

    anken_No = driver.find_element(By.XPATH,'//*[@id="ProposalNoText"]').get_attribute("value")#作成された案件番号を取得

    print(anken_No)
    #
    # //*[@id="ProposalNoText"]#案件番号のXPATH　　この中のvalueが案件番号

    anken_data = {'name':str(ym)+eplyName,'anken_No':str(anken_No)}
    #
    # anken_data.append(ym+eplyName,anken_No)
    #
    print(anken_data)

    driver.find_element(By.XPATH,'//*[@id="CloseButton"]/span').click()#閉じる釦（これをやらないとずっと更新中となり、案件削除できなくなるので注意）

    # sg.popup_ok(str(ym)+eplyName+'の案件番号'+anken_No,title = '案件番号')#ポップアップで案件番号表示。（案件番号は、一応取得済みなので見るだけ）


    # if ym == 202404:
    bk = pd.ExcelFile(r'c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/★勤務確認などのダウンロードデータ★/Emily_Files/Emily_anken.xlsx')


    wb = openpyxl.load_workbook('c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/★勤務確認などのダウンロードデータ★/Emily_Files/Emily_anken.xlsx')

    target_name =  str(ym)+'案件番号'


    check = False

    for ws in wb.worksheets:
        if ws.title == target_name:
            check = True

    if check == True:
        print(target_name + 'あるよ')
        df = pd.read_excel(bk, sheet_name=str(ym) + '案件番号', dtype=str, index_col=[0])

        df = pd.DataFrame(df)
        print(df)
        anken_data = {'name': str(ym) + eplyName, 'anken_No': str(anken_No)}
        print(anken_data)
        df_anken = pd.DataFrame(anken_data, index=[0])
        print(df_anken)
        df_new = pd.concat([df, df_anken])
        print(df_new)
        with pd.ExcelWriter(
                'c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/★勤務確認などのダウンロードデータ★/Emily_Files/Emily_anken.xlsx',
                engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_new.to_excel(writer, sheet_name=str(ym) + '案件番号', index=[0])
        # df_new.to_excel('c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/★勤務確認などのダウンロードデータ★/Emily_Files/Emily_anken.xlsx',sheet_name = str(ym)+'案件番号')
    else:
        print(target_name + 'ないよ')


        anken_data = {'name': str(ym) + eplyName, 'anken_No': str(anken_No)}
        print(anken_data)
        df_anken = pd.DataFrame(anken_data, index=[str(0)])
        print(df_anken)
        # df_new = pd.concat([df,df_anken])
        # print(df_new)
        with pd.ExcelWriter(
                'c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/★勤務確認などのダウンロードデータ★/Emily_Files/Emily_anken.xlsx',
                engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_anken.to_excel(writer, sheet_name=str(ym) + '案件番号', index=[0])

        # df_new.to_excel('c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/★勤務確認などのダウンロードデータ★/Emily_Files/Emily_anken.xlsx',sheet_name = str(ym)+'案件番号')


#====================エクセルファイルのセル幅自動調整===============================



    '''
        sheet_width.py
        purpose: make new xlsx and set width automatically
    '''

    import openpyxl as xl


    # set input file name
    inputfile = 'c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/★勤務確認などのダウンロードデータ★/Emily_Files/Emily_anken.xlsx'

    # read input xlsx
    wb1 = xl.load_workbook(filename=inputfile)

    print(wb1.sheetnames)
    s=len(wb1.sheetnames)

    for s in range(s):
        ws1 = wb1.worksheets[s]#[]内がワークシートの選択。ワークシート数を取得して、その数だけfor文でまわすとOKだと思う。

    # set column width
        for col in ws1.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            ws1.column_dimensions[col[0].column_letter].width = adjusted_width#column_dimension[column]ではダメ、左のように書き直し。openpyxl３以降では、intではだめで文字列。

    # save xlsx file
        wb1.save(inputfile)


#====================================================以上案径作成プログラム

    sg.theme('SystemDefault')
    layout = [[sg.Text('続けて別の人の案件作成しますか？', text_color='#FF0000', font=('meiryo,6'))],
              [sg.Button('はい', button_color=('red', '#808080'), key='-YES-'),
               sg.Button('いいえ', button_color=('blue', '#808080'), key='-NO-'),
               sg.Text('処理中断したいときは、Windowを閉じてください。', font=('Noto Serif CJK JP', 10))]]
    window = sg.Window('案件自動作成ツール', layout, size=(600, 150))
    event, values = window.read()
    if event == '-YES-':

        window.close()
        continue



    elif event == sg.WIN_CLOSED:
        break

    elif event == '-NO-':
        window.close()
        break

    window.close()

#sg.popup_ok(str(ym)+eplyName+'の案件番号'+anken_No,title = '案件番号')#ポップアップで案件番号表示。（案件番号は、一応取得済みなので見るだけ）

